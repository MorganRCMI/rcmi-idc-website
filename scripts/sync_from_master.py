#!/opt/homebrew/bin/python3
"""
sync_from_master.py

Reads PPP and SS pilot faculty + RCMI pilot publications from the master Excel
workbook and upserts them into rcmi_content.xlsx, then regenerates all CSVs.

Usage:
  python3 scripts/sync_from_master.py           # run for real
  python3 scripts/sync_from_master.py --dry-run # preview without writing
"""

import os
import re
import sys
import subprocess
import unicodedata
from collections import defaultdict

ROOT        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VENDOR_PATH = os.path.join(ROOT, "scripts", "_vendor")
if os.path.isdir(VENDOR_PATH) and VENDOR_PATH not in sys.path:
    sys.path.insert(0, VENDOR_PATH)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not found. Run: python3 -m pip install --target scripts/_vendor openpyxl", file=sys.stderr)
    raise SystemExit(1)

MASTER_PATH   = "/Users/mind/Library/CloudStorage/GoogleDrive-joain1@morgan.edu/Shared drives/Team REC IDC/Pilot Grants/02_Faculty_cleaned_wip.xlsx"
WORKBOOK_PATH = os.path.join(ROOT, "workbook", "rcmi_content.xlsx")
EXPORT_SCRIPT = os.path.join(ROOT, "scripts", "export_workbook_to_csv.py")
DRY_RUN       = "--dry-run" in sys.argv

# ── Column indices (0-based) in each master sheet ─────────────────────────

PPP_COL = dict(email=0, name=1, formal=2, program=4, status=5, rank=6,
               year=8, amount=11, category=12, project=13)

SS_COL  = dict(email=0, name=1, rank=3, program=4, year=5,
               status=7, amount=9, category=10, project=11)

NM_COL  = dict(email=0, formal=4, degree=5, rank=6, gender=7,
               dept=11, interest=12, gww=16, ppp=19, ss=22)

PUB_COL = dict(title=1, year=2, ptype=3, author=5, author_num=6,
               msu_dep=10, pilot_fpp=18)

# ── Helpers ────────────────────────────────────────────────────────────────

def slugify(text):
    s = unicodedata.normalize("NFKD", str(text or "").lower()).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")[:50]

def clean(val):
    """Normalize a cell value to a stripped string; blank out Excel errors."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.startswith("#") else s

def truthy(val):
    """Return True for any meaningful/non-empty cell value."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    return str(val).strip().lower() not in ("", "0", "no", "false", "-", "n/a", "na")

def read_rows(wb, sheet_name):
    return [tuple(c.value for c in row) for row in wb[sheet_name].iter_rows()]

PUB_TYPE_MAP = {
    "journal":    "Journal Article",
    "review":     "Review Article",
    "conference": "Conference Paper",
    "book":       "Book Chapter",
    "report":     "Report",
    "preprint":   "Preprint",
}

def map_pub_type(raw):
    low = str(raw or "").lower()
    for key, val in PUB_TYPE_MAP.items():
        if key in low:
            return val
    return "Journal Article"

# ── Read master workbook ───────────────────────────────────────────────────

print("Reading master workbook …")
master_wb = load_workbook(MASTER_PATH, read_only=True, data_only=True)

# NameMASTER: email (lowercase) → row tuple
name_master = {}
for row in read_rows(master_wb, "NameMASTER")[1:]:
    email = clean(row[NM_COL["email"]]).lower()
    if email:
        name_master[email] = row

# PPP rows — keep only "Funded" entries
ppp_rows = [r for r in read_rows(master_wb, "PPP")[1:]
            if clean(r[PPP_COL["status"]]).lower() == "funded"]

# SS rows — keep only "Funded" entries
ss_rows  = [r for r in read_rows(master_wb, "SS")[1:]
            if clean(r[SS_COL["status"]]).lower() == "funded"]

# RCMI-Pub — pilot publications have a non-null email in Pilot FPP (col 18)
pub_groups = defaultdict(list)
for row in read_rows(master_wb, "RCMI-Pub")[1:]:
    title = clean(row[PUB_COL["title"]])
    year  = row[PUB_COL["year"]]
    if title:
        pub_groups[(title, year)].append(row)

# A paper is "pilot" if any of its rows has a non-empty Pilot FPP value
pilot_pubs = {k: v for k, v in pub_groups.items()
              if any(truthy(r[PUB_COL["pilot_fpp"]]) for r in v)}

master_wb.close()

print(f"  PPP faculty: {len(ppp_rows)}")
print(f"  SS faculty:  {len(ss_rows)}")
print(f"  Pilot publications: {len(pilot_pubs)}")

# ── Build unified faculty map ──────────────────────────────────────────────
# Key: lowercase email → { programs: set, year: int, row data from each program }

faculty_map = {}  # email → dict with merged data

def add_to_faculty_map(email, formal, rank_val, year, amount, category, project, program_label):
    email = email.lower()
    if not email:
        return
    if email not in faculty_map:
        faculty_map[email] = {
            "email":    email,
            "formal":   formal,
            "rank":     rank_val,
            "programs": {},   # label → year
            "project":  project,
            "category": category,
            "amount":   amount,
        }
    entry = faculty_map[email]
    entry["programs"][program_label] = int(year) if truthy(year) else 9999
    # Keep the project from whichever program has the earliest year
    if not entry["project"] and project:
        entry["project"] = project

for row in ppp_rows:
    add_to_faculty_map(
        clean(row[PPP_COL["email"]]),
        clean(row[PPP_COL["formal"]]) or clean(row[PPP_COL["name"]]),
        clean(row[PPP_COL["rank"]]),
        row[PPP_COL["year"]],
        clean(row[PPP_COL["amount"]]),
        clean(row[PPP_COL["category"]]),
        clean(row[PPP_COL["project"]]),
        "PPP",
    )

for row in ss_rows:
    email   = clean(row[SS_COL["email"]])
    nm      = name_master.get(email.lower(), {})
    formal  = (clean(nm[NM_COL["formal"]]) if nm else "") or clean(row[SS_COL["name"]])
    add_to_faculty_map(
        email,
        formal,
        clean(row[SS_COL["rank"]]),
        row[SS_COL["year"]],
        clean(row[SS_COL["amount"]]),
        clean(row[SS_COL["category"]]),
        clean(row[SS_COL["project"]]),
        "SS",
    )

# ── Build faculty records sorted by earliest funded year ──────────────────

def build_program_type(entry, nm_row):
    """Combine program memberships from the entry + NameMASTER flags."""
    parts = []
    if nm_row and truthy(nm_row[NM_COL["gww"]]):
        parts.append("GWW")
    # Add programs in year order
    for label in sorted(entry["programs"], key=lambda p: entry["programs"][p]):
        if label not in parts:
            parts.append(label)
    # SS from NameMASTER (in case SS sheet missed someone)
    if nm_row and truthy(nm_row[NM_COL["ss"]]) and "SS" not in parts:
        parts.append("SS")
    return ", ".join(parts)

# Sort by earliest year then name for stable ordering
sorted_faculty = sorted(
    faculty_map.values(),
    key=lambda e: (min(e["programs"].values()), (e["formal"] or e["email"]).lower()),
)

new_faculty = []
for idx, entry in enumerate(sorted_faculty):
    email   = entry["email"]
    nm      = name_master.get(email)
    formal  = entry["formal"] or (clean(nm[NM_COL["formal"]]) if nm else email)
    dept    = clean(nm[NM_COL["dept"]])     if nm else ""
    interest = clean(nm[NM_COL["interest"]]) if nm else ""
    degree  = clean(nm[NM_COL["degree"]])   if nm else ""
    gender  = clean(nm[NM_COL["gender"]]).lower() if nm else ""
    rank    = entry["rank"] or (clean(nm[NM_COL["rank"]]) if nm else "")
    icon    = "👩‍🔬" if "female" in gender else "👨‍🔬"

    earliest_year = min(entry["programs"].values())
    program_type  = build_program_type(entry, nm)
    ppp_year      = entry["programs"].get("PPP")
    ss_year       = entry["programs"].get("SS")
    notes_parts   = []
    if ppp_year and ppp_year != 9999:
        notes_parts.append(f"PPP {ppp_year}")
    if ss_year and ss_year != 9999:
        notes_parts.append(f"SS {ss_year}")
    if entry["amount"]:
        notes_parts.append(f"${entry['amount']}")

    new_faculty.append({
        "Faculty ID":         "pilot_" + slugify(email.split("@")[0]),
        "Is Active":          "Yes",
        "Sort Order":         earliest_year * 1000 + idx,
        "Full Name":          formal,
        "Category":           "Pilot Faculty",
        "Title":              rank,
        "Department":         dept,
        "Summary Label":      "Research Interests",
        "Summary Text":       interest,
        "Email":              email,
        "Fallback Icon":      icon,
        "Education 1":        degree,
        "Education 2":        "",
        "Education 3":        "",
        "Education 4":        "",
        "Tag 1":              entry.get("category", ""),
        "Tag 2":              "",
        "Tag 3":              "",
        "Tag 4":              "",
        "Tag 5":              "",
        "Tag 6":              "",
        "Highlight Heading":  "Pilot Project",
        "Highlight Text":     entry.get("project", ""),
        "Office":             "",
        "Phone":              "",
        "Image Path":         "",
        "Image Alt Text":     formal,
        "Internal Notes":     " | ".join(notes_parts),
        "Year Funded":        earliest_year if earliest_year != 9999 else "",
        "Program Type":       program_type,
    })

# ── Build publication records ──────────────────────────────────────────────

# Build a quick lookup: faculty email → program type string
email_to_program = {e["email"]: build_program_type(e, name_master.get(e["email"]))
                    for e in faculty_map.values()}

new_publications = []
for sort_idx, ((title, year), rows) in enumerate(
        sorted(pilot_pubs.items(), key=lambda x: -(int(x[0][1]) if x[0][1] else 0))):

    # Sort authors by author number
    try:
        sorted_rows = sorted(rows, key=lambda r: int(r[PUB_COL["author_num"]]) if truthy(r[PUB_COL["author_num"]]) else 999)
    except (TypeError, ValueError):
        sorted_rows = rows

    authors, seen_a = [], set()
    msu_dep, ptype_raw = "", ""
    pilot_email = ""

    for r in sorted_rows:
        a = clean(r[PUB_COL["author"]])
        if a and a.lower() not in seen_a:
            seen_a.add(a.lower())
            authors.append(a)
        if not msu_dep and clean(r[PUB_COL["msu_dep"]]):
            msu_dep = clean(r[PUB_COL["msu_dep"]])
        if not ptype_raw and clean(r[PUB_COL["ptype"]]):
            ptype_raw = clean(r[PUB_COL["ptype"]])
        if not pilot_email and truthy(r[PUB_COL["pilot_fpp"]]):
            pilot_email = clean(r[PUB_COL["pilot_fpp"]]).lower()

    # Determine program type from the pilot faculty's programs
    prog_type = email_to_program.get(pilot_email, "PPP")

    new_publications.append({
        "Publication ID":          "rcmi_pub_" + slugify(title[:30]) + "_" + str(year or ""),
        "Project ID":              "",
        "Is Active":               "Yes",
        "Sort Order":              sort_idx + 100,
        "Title":                   title,
        "Authors":                 "; ".join(authors),
        "Authors Short":           authors[0] + (" et al." if len(authors) > 1 else "") if authors else "",
        "Year":                    str(year or ""),
        "Publication Type":        map_pub_type(ptype_raw),
        "Department":              msu_dep,
        "Journal or Source":       "",
        "Citation Text":           "",
        "DOI":                     "",
        "Abstract":                "",
        "Featured Label":          "",
        "Full Text URL":           "",
        "Project Display Override": "",
        "Internal Notes":          f"Pilot PI: {pilot_email}" if pilot_email else "",
        "Program Type":            prog_type,
    })

print(f"  → {len(new_faculty)} faculty records to upsert")
print(f"  → {len(new_publications)} publication records to upsert")

if DRY_RUN:
    print("\n[DRY RUN] No changes written. Sample faculty:")
    for rec in new_faculty[:3]:
        print(f"  {rec['Full Name']} | {rec['Program Type']} | Year: {rec['Year Funded']}")
    print("Sample publications:")
    for rec in new_publications[:3]:
        print(f"  {rec['Title'][:60]} ({rec['Year']}) | {rec['Program Type']}")
    raise SystemExit(0)

# ── Upsert into workbook ───────────────────────────────────────────────────

def get_or_add_col(ws, headers, col_name):
    """Return 1-based column index for col_name, adding it as a new header if missing."""
    if col_name in headers:
        return headers.index(col_name) + 1
    new_idx = len(headers) + 1
    headers.append(col_name)
    ws.cell(row=1, column=new_idx).value = col_name
    return new_idx

def upsert_sheet(wb, sheet_name, records, key_col):
    ws      = wb[sheet_name]
    headers = [str(c.value or "") for c in ws[1]]

    # Ensure all required columns exist
    for record in records:
        for col_name in record:
            get_or_add_col(ws, headers, col_name)

    key_col_idx = headers.index(key_col)  # 0-based for reading

    # Detect a "Category" column for faculty-sheet protection logic
    cat_col_idx = headers.index("Category") + 1 if "Category" in headers else None

    # Index existing rows by key value (lowercase for case-insensitive match)
    existing = {}
    for row_num in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_num, column=key_col_idx + 1).value
        if cell_val:
            existing[str(cell_val).lower().strip()] = row_num

    # Fields we are allowed to add/update on an already-existing, non-pilot row
    PILOT_METADATA = {"Year Funded", "Program Type"}

    added = updated = 0
    for record in records:
        key_val = str(record.get(key_col, "")).lower().strip()
        row_num = existing.get(key_val)

        if row_num:
            # If an existing row has a non-Pilot category, only backfill the
            # pilot metadata columns — leave everything else (name, category,
            # title, dept…) exactly as the editor set it.
            if cat_col_idx:
                existing_cat = str(ws.cell(row=row_num, column=cat_col_idx).value or "")
                if existing_cat and existing_cat != "Pilot Faculty":
                    for col_name in PILOT_METADATA:
                        if col_name in record:
                            ws.cell(row=row_num,
                                    column=get_or_add_col(ws, headers, col_name)
                                    ).value = record[col_name]
                    updated += 1
                    continue   # skip full overwrite

            updated += 1
        else:
            row_num = ws.max_row + 1
            existing[key_val] = row_num
            added += 1

        for col_name, value in record.items():
            col_idx = get_or_add_col(ws, headers, col_name)
            ws.cell(row=row_num, column=col_idx).value = value

    return added, updated

print("\nOpening workbook for writing …")
content_wb = load_workbook(WORKBOOK_PATH)

fac_added,  fac_upd  = upsert_sheet(content_wb, "faculty",      new_faculty,       "Email")
pub_added,  pub_upd  = upsert_sheet(content_wb, "publications",  new_publications,  "Title")

content_wb.save(WORKBOOK_PATH)
print(f"  faculty:      {fac_added} added, {fac_upd} updated")
print(f"  publications: {pub_added} added, {pub_upd} updated")

# ── Regenerate CSVs ────────────────────────────────────────────────────────
print("\nRegenerating CSVs …")
subprocess.run([sys.executable, EXPORT_SCRIPT], check=True)
print("\nAll done.")
