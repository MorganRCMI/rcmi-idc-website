#!/Users/mind/jupyter_env/bin/python3
"""
fetch_faculty_photos.py

Fetches faculty headshots from Morgan State profile pages and saves them
to docs/img/faculty/. Updates Image Path in rcmi_content.xlsx, re-exports CSVs.

Strategy (in order):
  1. Direct URL construction: /dept-slug/faculty-and-staff/name-slug
  2. Department scan: search all known MSU dept listing pages for the name
  3. Bing search fallback: decode real URLs from Bing redirect params

Usage:
  ~/jupyter_env/bin/python3 scripts/fetch_faculty_photos.py           # all missing photos
  ~/jupyter_env/bin/python3 scripts/fetch_faculty_photos.py --dry-run
  ~/jupyter_env/bin/python3 scripts/fetch_faculty_photos.py --all     # re-fetch even existing
  ~/jupyter_env/bin/python3 scripts/fetch_faculty_photos.py --all-categories  # include non-pilot
"""

import os
import re
import sys
import time
import base64
import subprocess
import unicodedata
import urllib.parse
from pathlib import Path

ROOT        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VENDOR_PATH = os.path.join(ROOT, "scripts", "_vendor")
if os.path.isdir(VENDOR_PATH) and VENDOR_PATH not in sys.path:
    sys.path.insert(0, VENDOR_PATH)

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    print("Error: run  pip install requests beautifulsoup4", file=sys.stderr)
    raise SystemExit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not found.", file=sys.stderr)
    raise SystemExit(1)

WORKBOOK_PATH = os.path.join(ROOT, "workbook", "rcmi_content.xlsx")
PHOTO_DIR     = os.path.join(ROOT, "docs", "img", "faculty")
EXPORT_SCRIPT = os.path.join(ROOT, "scripts", "export_workbook_to_csv.py")
LOG_PATH      = os.path.join(ROOT, "scripts", "photo_fetch_log.txt")
MSU_BASE      = "https://www.morgan.edu"

DRY_RUN       = "--dry-run"       in sys.argv
REFETCH       = "--all"           in sys.argv
ALL_CATS      = "--all-categories" in sys.argv

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
})

os.makedirs(PHOTO_DIR, exist_ok=True)

# ── All known MSU department slugs with faculty-and-staff pages ──────────────

ALL_DEPT_SLUGS = [
    "biology", "chemistry", "mathematics",
    "physics-engineering-physics",
    "computer-science",
    "electrical-computer-engineering",
    "industrial-and-systems-engineering",
    "civil-engineering",
    "mechatronics-engineering",
    "psychology",
    "school-of-community-health-and-policy",
    "social-work",
    "public-health",
    "accounting-and-finance",
    "medical-laboratory-science",
    "architecture-and-planning",
    "education",
    "school-of-education-and-urban-studies",
    "advanced-studies-leadership-and-policy",
    "nursing",
    "engineering",
    "research-and-economic-development",
    "school-of-business-and-management",
]

# Cache of dept listing pages: slug → set of profile hrefs
_dept_cache: dict[str, set] = {}

def _load_dept_page(dept_slug):
    if dept_slug in _dept_cache:
        return _dept_cache[dept_slug]
    url = f"{MSU_BASE}/{dept_slug}/faculty-and-staff"
    try:
        r = SESSION.get(url, timeout=8)
        if r.status_code != 200:
            _dept_cache[dept_slug] = set()
            return set()
        soup = BeautifulSoup(r.text, "html.parser")
        hrefs = set()
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if f"/{dept_slug}/faculty-and-staff/" in href and href != f"/{dept_slug}/faculty-and-staff":
                hrefs.add(href)
        _dept_cache[dept_slug] = hrefs
        return hrefs
    except Exception:
        _dept_cache[dept_slug] = set()
        return set()

# ── Helpers ──────────────────────────────────────────────────────────────────

def log(msg):
    print(msg)
    with open(LOG_PATH, "a") as f:
        f.write(msg + "\n")

def slugify(text):
    s = unicodedata.normalize("NFKD", str(text or "").lower()).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")

def name_to_slug(full_name):
    """
    'Dr. Ingrid K. Tulloch, Ph.D.' → 'ingrid-tulloch'
    Strips leading titles, middle initials, trailing credentials.
    """
    name = re.sub(r"^(dr\.?|prof\.?|mr\.?|ms\.?|mrs\.?)\s+", "", full_name.strip(), flags=re.I)
    name = re.sub(r",.*$", "", name)                        # drop ", Ph.D." etc.
    name = re.sub(r"\s+(jr\.?|sr\.?|ii|iii)$", "", name, flags=re.I)
    # Strip single-letter middle initial (e.g. "K." between first and last)
    name = re.sub(r"\s+[A-Z]\.\s+", " ", name)
    return slugify(name.strip())

def dept_to_slug(dept):
    """
    Convert a department label to candidate Morgan State URL slugs.
    Returns list, most specific first.
    """
    cleaned = re.sub(r"^(department of|school of|division of|college of)\s+", "", dept.strip(), flags=re.I)
    base = slugify(cleaned)

    known = {
        "community-health-policy":              ["school-of-community-health-and-policy"],
        "community-health-and-policy":          ["school-of-community-health-and-policy"],
        "electrical-and-computer-engineering":  ["electrical-computer-engineering", "engineering"],
        "electrical-computer-engineering":      ["electrical-computer-engineering", "engineering"],
        "civil-engineering":                    ["civil-engineering", "engineering"],
        "industrial-systems-engineering":       ["industrial-and-systems-engineering", "engineering"],
        "industrial-and-systems-engineering":   ["industrial-and-systems-engineering", "engineering"],
        "computer-science":                     ["computer-science"],
        "physics-and-engineering-physics":      ["physics-engineering-physics", "physics"],
        "physics-engineering-physics":          ["physics-engineering-physics"],
        "mathematics":                          ["mathematics"],
        "biology":                              ["biology"],
        "chemistry":                            ["chemistry"],
        "psychology":                           ["psychology"],
        "public-health":                        ["public-health", "school-of-community-health-and-policy"],
        "social-work":                          ["social-work"],
        "accounting-and-finance":               ["accounting-and-finance", "school-of-business-and-management"],
        "accounting":                           ["accounting-and-finance", "school-of-business-and-management"],
        "medical-laboratory-science":           ["medical-laboratory-science"],
        "clinical-laboratory-science":          ["medical-laboratory-science"],
        "architecture-and-planning":            ["architecture-and-planning"],
        "architecture":                         ["architecture-and-planning"],
        "mechatronics-engineering":             ["mechatronics-engineering", "industrial-and-systems-engineering", "engineering"],
        "mechatronics":                         ["mechatronics-engineering", "engineering"],
        "education":                            ["education", "school-of-education-and-urban-studies"],
        "advanced-studies-leadership-and-policy": ["advanced-studies-leadership-and-policy", "education"],
        "advances-studies-leadership-and-policy": ["advanced-studies-leadership-and-policy"],
        "research-and-economic-development":    ["research-and-economic-development"],
        "research":                             ["research-and-economic-development"],
        "nursing":                              ["nursing"],
    }

    for key, vals in known.items():
        if key == base or key in base or base in key:
            return vals + ([base] if base not in vals else [])

    return [base]

def extract_photo_from_page(url):
    """Scrape MSU profile page and return absolute photo URL or None."""
    try:
        r = SESSION.get(url, timeout=10)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "html.parser")
        img  = soup.find("img", class_="lazyload")
        if not img:
            last = url.rstrip("/").split("/")[-1].split("-")[-1]
            img  = soup.find("img", alt=re.compile(last, re.I))
        if img:
            src = img.get("data-src") or img.get("src") or ""
            if src and not src.startswith("data:") and "logo" not in src.lower():
                return MSU_BASE + src if src.startswith("/") else src
    except Exception:
        pass
    return None

# ── Strategy 1: direct URL construction ─────────────────────────────────────

def fetch_direct(full_name, department):
    name_slug  = name_to_slug(full_name)
    dept_slugs = dept_to_slug(department) if department.strip() else []
    for dept_slug in dept_slugs:
        url = f"{MSU_BASE}/{dept_slug}/faculty-and-staff/{name_slug}"
        photo = extract_photo_from_page(url)
        if photo:
            return photo
    return None

# ── Strategy 2: scan all dept listing pages ──────────────────────────────────

def fetch_by_dept_scan(full_name):
    """Search each known dept listing page for a link matching the name slug."""
    name_slug = name_to_slug(full_name)
    for dept_slug in ALL_DEPT_SLUGS:
        hrefs = _load_dept_page(dept_slug)
        for href in hrefs:
            link_slug = href.rstrip("/").split("/")[-1]
            if link_slug == name_slug:
                url   = MSU_BASE + href
                photo = extract_photo_from_page(url)
                if photo:
                    return photo, href
        time.sleep(0.1)
    return None, None

# ── Strategy 3: Bing search fallback ────────────────────────────────────────

def _decode_bing_url(href):
    parsed = urllib.parse.urlparse(href)
    params = urllib.parse.parse_qs(parsed.query)
    u = params.get("u", [""])[0]
    if u.startswith("a1"):
        b64 = u[2:] + "=" * (-len(u[2:]) % 4)
        try:
            return base64.urlsafe_b64decode(b64).decode("utf-8", errors="replace")
        except Exception:
            pass
    return None

def fetch_via_web_search(full_name):
    """Search Bing for '[name] morgan state university', find faculty profile, scrape photo."""
    query = f"{full_name} morgan state university"
    try:
        r = SESSION.get(
            f"https://www.bing.com/search?q={urllib.parse.quote(query)}",
            timeout=10
        )
        soup = BeautifulSoup(r.text, "html.parser")

        # Look for morgan.edu/faculty-and-staff links first
        for a in soup.select("li.b_algo h2 a"):
            href = a.get("href", "")
            real = _decode_bing_url(href)
            if real and "morgan.edu" in real and "/faculty-and-staff/" in real:
                photo = extract_photo_from_page(real)
                if photo:
                    return photo, real

        # Fallback: try any morgan.edu link
        for a in soup.select("li.b_algo h2 a"):
            href = a.get("href", "")
            real = _decode_bing_url(href)
            if real and "morgan.edu" in real:
                photo = extract_photo_from_page(real)
                if photo:
                    return photo, real
    except Exception:
        pass
    return None, None

# ── Orchestrate ──────────────────────────────────────────────────────────────

def get_photo(full_name, department):
    """Try all strategies; return (photo_url, method_label) or (None, None)."""
    # 1. Direct URL (if department provided)
    if department.strip():
        photo = fetch_direct(full_name, department)
        if photo:
            return photo, "direct"

    # 2. Web search (quick, works for most cases)
    time.sleep(0.3)  # polite delay
    photo, profile_url = fetch_via_web_search(full_name)
    if photo:
        return photo, f"search:{profile_url}"

    # 3. Dept listing scan (fallback, slower but thorough)
    photo, profile_href = fetch_by_dept_scan(full_name)
    if photo:
        return photo, f"scan:{profile_href}"

    return None, None

def download_photo(photo_url, save_path):
    try:
        resp = SESSION.get(photo_url, timeout=15)
        resp.raise_for_status()
        ct = resp.headers.get("Content-Type", "")
        if not any(t in ct for t in ("image/", "jpeg", "png", "webp")):
            return False
        with open(save_path, "wb") as f:
            f.write(resp.content)
        return True
    except Exception:
        return False

# ── Load faculty needing photos ───────────────────────────────────────────────

print("Loading workbook …")
wb = load_workbook(WORKBOOK_PATH)
ws = wb["faculty"]
headers = [str(c.value or "") for c in ws[1]]

def col(name):
    return headers.index(name) + 1

try:
    id_col   = col("Faculty ID")
    name_col = col("Full Name")
    dept_col = col("Department")
    cat_col  = col("Category")
    img_col  = col("Image Path")
except ValueError as exc:
    print(f"Error: missing column {exc}", file=sys.stderr)
    raise SystemExit(1)

PILOT_ONLY_CATS = {"Pilot Faculty"} if not ALL_CATS else None

targets = []
for row_num in range(2, ws.max_row + 1):
    category = str(ws.cell(row=row_num, column=cat_col).value or "")
    img_path = str(ws.cell(row=row_num, column=img_col).value or "").strip()
    if PILOT_ONLY_CATS and category not in PILOT_ONLY_CATS:
        continue
    # Skip if already has a valid relative image path
    if img_path and img_path.startswith("./img/") and not REFETCH:
        continue
    targets.append({
        "row":  row_num,
        "id":   str(ws.cell(row=row_num, column=id_col).value or ""),
        "name": str(ws.cell(row=row_num, column=name_col).value or ""),
        "dept": str(ws.cell(row=row_num, column=dept_col).value or ""),
    })

print(f"  {len(targets)} faculty need photos")

if DRY_RUN:
    for t in targets[:10]:
        slugs = dept_to_slug(t["dept"]) if t["dept"].strip() else ["(no dept — will scan)"]
        print(f"  {t['name']}  →  /{slugs[0]}/faculty-and-staff/{name_to_slug(t['name'])}")
    print("[DRY RUN] No downloads performed.")
    raise SystemExit(0)

# ── Fetch ─────────────────────────────────────────────────────────────────────

with open(LOG_PATH, "a") as f:
    f.write("\n=== fetch run ===\n")

found = failed = 0

for target in targets:
    name = target["name"]
    dept = target["dept"]
    fid  = target["id"]

    print(f"  {name} ({dept or 'no dept'}) … ", end="", flush=True)
    time.sleep(0.4)

    photo_url, method = get_photo(name, dept)

    if not photo_url:
        print("not found")
        log(f"FAIL | {name} | {dept}")
        failed += 1
        continue

    ext       = ".jpg"
    url_lower = urllib.parse.urlparse(photo_url).path.lower()
    for candidate in (".png", ".webp", ".jpeg"):
        if url_lower.endswith(candidate):
            ext = candidate
            break

    save_path = os.path.join(PHOTO_DIR, fid + ext)
    if download_photo(photo_url, save_path):
        rel = f"./img/faculty/{fid}{ext}"
        ws.cell(row=target["row"], column=img_col).value = rel
        print(f"saved [{method}]")
        log(f"OK  | {name} | {rel} | {method}")
        found += 1
    else:
        print("download failed")
        log(f"FAIL (download) | {name} | {photo_url}")
        failed += 1

print(f"\n  Found: {found}  |  Not found: {failed}")
print(f"  Log: {os.path.relpath(LOG_PATH, ROOT)}")

wb.save(WORKBOOK_PATH)
print("\nRegenerating CSVs …")
subprocess.run([sys.executable, EXPORT_SCRIPT], check=True)
print("Done.")
