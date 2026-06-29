#!/usr/bin/env python3
"""
sync_from_ppp.py

Syncs faculty, research, and publications data from rcmi_ppp_data_tmp.xlsx
(single source of truth) into rcmi_content.xlsx.

Strategy:
  - Keep all existing Leadership/PI rows in faculty sheet
  - Replace Pilot Faculty rows with 33 from PPP sheet
  - Extract research projects from "Project Title / Area" in PPP
  - Extract publications from RCMI_publications sheet

Usage:
  python3 scripts/sync_from_ppp.py
"""

import os
import sys
import shutil
import re
import unicodedata
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VENDOR = os.path.join(ROOT, "scripts", "_vendor")
if os.path.isdir(VENDOR) and VENDOR not in sys.path:
    sys.path.insert(0, VENDOR)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not found", file=sys.stderr)
    raise SystemExit(1)

PPP_PATH = os.path.join(ROOT, "workbook", "rcmi_ppp_data_tmp.xlsx")
CONTENT_PATH = os.path.join(ROOT, "workbook", "rcmi_content.xlsx")
BACKUP_PATH = os.path.join(ROOT, "workbook", "rcmi_content_backup_before_ppp_sync.xlsx")
EXPORT_SCRIPT = os.path.join(ROOT, "scripts", "export_workbook_to_csv.py")

def clean(v):
    return str(v or "").strip()

def slugify(text):
    s = unicodedata.normalize("NFKD", str(text or "").lower()).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")

def name_to_slug(full_name):
    name = re.sub(r"^(dr\.?|prof\.?|mr\.?|ms\.?|mrs\.?)\s+", "", full_name.strip(), flags=re.I)
    name = re.sub(r",.*$", "", name)
    name = re.sub(r"\s+(jr\.?|sr\.?|ii|iii)$", "", name, flags=re.I)
    return slugify(name.strip())

# ── Load source data ──────────────────────────────────────────────────────────

print("Loading source data from rcmi_ppp_data_tmp.xlsx ...")
ppp_wb = load_workbook(PPP_PATH, data_only=True, read_only=True)

ppp_headers = [clean(c.value) for c in ppp_wb["PPP"][1]]
ppp_rows = [
    tuple(c.value for c in row)
    for row in ppp_wb["PPP"].iter_rows(min_row=2, values_only=False)
    if any(c.value for c in row)
]

pub_headers = [clean(c.value) for c in ppp_wb["RCMI_publications"][1]]
pub_rows = [
    tuple(c.value for c in row)
    for row in ppp_wb["RCMI_publications"].iter_rows(min_row=2, values_only=False)
    if any(c.value for c in row)
]

print(f"  PPP rows: {len(ppp_rows)}")
print(f"  Publications rows: {len(pub_rows)}")

def col(headers, name):
    try:
        return headers.index(name)
    except ValueError:
        return None

# Column indices for PPP
ppp_cols = {
    "email": col(ppp_headers, "email (Key ID)"),
    "name": col(ppp_headers, "Name"),
    "formal": col(ppp_headers, "Formal name"),
    "fid": col(ppp_headers, "facultyfundingID"),
    "rank": col(ppp_headers, "rank"),
    "gww": col(ppp_headers, "GWW?"),
    "year": col(ppp_headers, "PPP Year"),
    "category": col(ppp_headers, "Category"),
    "project": col(ppp_headers, "Project Title / Area"),
    "publications": col(ppp_headers, "publications"),
}

# Column indices for publications
pub_cols = {
    "email": col(pub_headers, "MSU-Email1"),
    "title": col(pub_headers, "TITLE"),
    "year": col(pub_headers, "YEAR"),
    "authors": col(pub_headers, "Author Name"),
    "department": col(pub_headers, "MSU Dep"),
    "category": col(pub_headers, "Category"),
}

print(f"\nPPP columns found: {len([v for v in ppp_cols.values() if v is not None])}/{len(ppp_cols)}")
print(f"Publication columns found: {len([v for v in pub_cols.values() if v is not None])}/{len(pub_cols)}")

# Validate critical columns
for key, idx in ppp_cols.items():
    if idx is None:
        print(f"ERROR: Missing PPP column '{key}'", file=sys.stderr)
        raise SystemExit(1)

# ── Parse faculty from PPP ────────────────────────────────────────────────────

print("\nProcessing PPP faculty ...")
faculty_records = []

for row in ppp_rows:
    email = clean(row[ppp_cols["email"]])
    formal = clean(row[ppp_cols["formal"]])
    rank = clean(row[ppp_cols["rank"]])
    gww = row[ppp_cols["gww"]]
    year = row[ppp_cols["year"]]
    category = clean(row[ppp_cols["category"]])
    project = clean(row[ppp_cols["project"]])

    if not email:
        continue

    # Program type logic: if GWW?=1 then "PPP-GWW" else "PPP"
    program_type = "PPP-GWW" if (gww == 1 or str(gww).strip() == "1") else "PPP"

    fid = f"pilot_{name_to_slug(formal or email)}"

    faculty_records.append({
        "Faculty ID": fid,
        "Full Name": formal,
        "Email": email,
        "Title": rank if rank and rank != "0" else "",
        "Department": category,
        "Summary Text": project,
        "Year Funded": year if year else None,
        "Program Type": program_type,
        "Category": "Pilot Faculty",
        "Is Active": "Yes",
    })

print(f"  Extracted {len(faculty_records)} pilot faculty")

# ── Parse research from PPP ───────────────────────────────────────────────────

print("Processing PPP research projects ...")
research_records = {}

for row in ppp_rows:
    project_title = clean(row[ppp_cols["project"]])
    email = clean(row[ppp_cols["email"]])
    formal = clean(row[ppp_cols["formal"]])
    category = clean(row[ppp_cols["category"]])

    if not project_title:
        continue

    # One project per unique title (de-duplicate if same project appears multiple times)
    if project_title not in research_records:
        pi_fac_id = f"pilot_{name_to_slug(formal)}" if formal else ""
        research_records[project_title] = {
            "Title": project_title,
            "Summary": project_title,
            "Project ID": f"ppp_{slugify(project_title[:40])}",
            "Row Type": "project",
            "Is Active": 1,
            "Category": category if category else "Research",
            "PI Faculty ID": pi_fac_id,
            "PI Name": formal,
            "Department": category,
        }

print(f"  Extracted {len(research_records)} research projects")

# ── Build faculty project map (email -> project_id) ────────────────────────────

faculty_to_project = {}
for row in ppp_rows:
    email = clean(row[ppp_cols["email"]])
    formal = clean(row[ppp_cols["formal"]])
    project_title = clean(row[ppp_cols["project"]])

    if email and project_title:
        project_id = f"ppp_{slugify(project_title[:40])}"
        faculty_to_project[email.lower()] = project_id

# ── Parse publications from RCMI_publications ─────────────────────────────────

print("Processing publications ...")
pub_by_key = defaultdict(list)

for row in pub_rows:
    title = clean(row[pub_cols["title"]])
    year = row[pub_cols["year"]]
    authors = clean(row[pub_cols["authors"]])
    email = clean(row[pub_cols["email"]])
    dept = clean(row[pub_cols["department"]])
    category = clean(row[pub_cols["category"]])

    if not title:
        continue

    key = (title, year)
    pub_by_key[key].append({
        "author": authors,
        "email": email,
        "department": dept,
        "category": category,
    })

publication_records = []
for (title, year), authors_list in pub_by_key.items():
    # Aggregate authors (unique)
    author_names = [a["author"] for a in authors_list if a["author"]]
    author_str = ", ".join(dict.fromkeys(author_names))  # deduplicate while preserving order
    dept = authors_list[0]["department"] if authors_list else ""
    cat = authors_list[0]["category"] if authors_list else ""

    # Try to match to a project by email
    project_id = None
    for item in authors_list:
        email = item.get("email", "").lower().strip()
        if email in faculty_to_project:
            project_id = faculty_to_project[email]
            break

    publication_records.append({
        "Title": title,
        "Year": year,
        "Authors": author_str,
        "Department": dept,
        "Publication Type": cat if cat else "Journal Article",
        "Is Active": "Yes",
        "Program Type": "PPP",  # All from this sheet are PPP
        "Project ID": project_id,  # Will be None if no match found
    })

print(f"  Extracted {len(publication_records)} publications")
matched = sum(1 for p in publication_records if p.get("Project ID"))
unmatched = len(publication_records) - matched
print(f"  Matched to projects: {matched}  |  Unmatched: {unmatched}")

# ── Load and update rcmi_content.xlsx ─────────────────────────────────────────

print("\nLoading rcmi_content.xlsx ...")
content_wb = load_workbook(CONTENT_PATH)

# Backup before modifying
print(f"Backing up to {os.path.relpath(BACKUP_PATH, ROOT)} ...")
shutil.copy(CONTENT_PATH, BACKUP_PATH)

# ── Update faculty sheet ──────────────────────────────────────────────────────

print("\nUpdating faculty sheet ...")
fac_ws = content_wb["faculty"]
fac_hdrs = [clean(c.value) for c in fac_ws[1]]

# Find and delete all "Pilot Faculty" rows (keep Leadership/PI)
pilot_rows_to_delete = []
for row_num in range(2, fac_ws.max_row + 1):
    cat = clean(fac_ws.cell(row=row_num, column=fac_hdrs.index("Category") + 1).value or "")
    if cat == "Pilot Faculty":
        pilot_rows_to_delete.append(row_num)

# Delete in reverse order to preserve row numbers
for row_num in reversed(pilot_rows_to_delete):
    fac_ws.delete_rows(row_num)

print(f"  Deleted {len(pilot_rows_to_delete)} existing Pilot Faculty rows")

# Add new faculty records
def col_idx(headers, name):
    try:
        return headers.index(name) + 1
    except ValueError:
        return None

fac_col = {k: col_idx(fac_hdrs, k) for k in ["Faculty ID", "Full Name", "Email", "Title", "Department", "Summary Text", "Year Funded", "Program Type", "Category", "Is Active", "Sort Order"]}

# Calculate sort orders (starting at 100, incrementing by 1)
sort_order = 100
for rec in faculty_records:
    rec["Sort Order"] = sort_order
    sort_order += 1
    rec["Is Active"] = "Yes"  # Text, not number
    row_num = fac_ws.max_row + 1
    for key, col_num in fac_col.items():
        if col_num:
            fac_ws.cell(row=row_num, column=col_num).value = rec.get(key)
    # Also set Sort Order
    sort_col = fac_col.get("Sort Order")
    if not sort_col:
        sort_col = fac_hdrs.index("Sort Order") + 1
    fac_ws.cell(row=row_num, column=sort_col).value = rec.get("Sort Order")

print(f"  Added {len(faculty_records)} new Pilot Faculty rows")

# ── Update research sheet ─────────────────────────────────────────────────────

print("\nUpdating research sheet ...")
res_ws = content_wb["research"]
res_hdrs = [clean(c.value) for c in res_ws[1]]

# Delete all PPP projects (keep existing infrastructure/other)
ppp_rows_to_delete = []
for row_num in range(2, res_ws.max_row + 1):
    proj_id = clean(res_ws.cell(row=row_num, column=res_hdrs.index("Project ID") + 1).value or "")
    if proj_id.startswith("ppp_"):
        ppp_rows_to_delete.append(row_num)

for row_num in reversed(ppp_rows_to_delete):
    res_ws.delete_rows(row_num)

print(f"  Deleted {len(ppp_rows_to_delete)} existing PPP project rows")

# Add research records
res_col = {k: col_idx(res_hdrs, k) for k in ["Row Type", "Project ID", "Title", "Summary", "Is Active", "Category", "PI Faculty ID", "PI Name", "Department"]}

for rec in research_records.values():
    row_num = res_ws.max_row + 1
    for key, col_num in res_col.items():
        if col_num:
            res_ws.cell(row=row_num, column=col_num).value = rec.get(key)

print(f"  Added {len(research_records)} new research projects")

# ── Update publications sheet ─────────────────────────────────────────────────

print("\nUpdating publications sheet ...")
pub_ws = content_wb["publications"]
pub_hdrs = [clean(c.value) for c in pub_ws[1]]

# Delete all PPP publications (keep existing)
ppp_pub_rows = []
for row_num in range(2, pub_ws.max_row + 1):
    prog = clean(pub_ws.cell(row=row_num, column=pub_hdrs.index("Program Type") + 1).value or "")
    if prog == "PPP":
        ppp_pub_rows.append(row_num)

for row_num in reversed(ppp_pub_rows):
    pub_ws.delete_rows(row_num)

print(f"  Deleted {len(ppp_pub_rows)} existing PPP publication rows")

# Add publication records
pub_col = {k: col_idx(pub_hdrs, k) for k in ["Title", "Year", "Authors", "Department", "Publication Type", "Program Type", "Is Active", "Project ID"]}

for rec in publication_records:
    row_num = pub_ws.max_row + 1
    for key, col_num in pub_col.items():
        if col_num:
            pub_ws.cell(row=row_num, column=col_num).value = rec.get(key)

print(f"  Added {len(publication_records)} new publications")

# ── Save and export ───────────────────────────────────────────────────────────

print("\nSaving rcmi_content.xlsx ...")
content_wb.save(CONTENT_PATH)

print("Regenerating CSV exports ...")
import subprocess
subprocess.run([sys.executable, EXPORT_SCRIPT], check=True)

print("\n" + "="*70)
print("SYNC COMPLETE")
print("="*70)
print(f"Faculty:      {len(faculty_records)} pilot faculty rows added")
print(f"Research:     {len(research_records)} project rows added")
print(f"Publications: {len(publication_records)} publication rows added")
print(f"\nBackup saved: {os.path.relpath(BACKUP_PATH, ROOT)}")
print("="*70)
