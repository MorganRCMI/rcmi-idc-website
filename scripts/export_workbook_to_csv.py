#!/opt/homebrew/bin/python3
import csv
import os
import sys
import tempfile


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VENDOR_PATH = os.path.join(ROOT, "scripts", "_vendor")

if os.path.isdir(VENDOR_PATH) and VENDOR_PATH not in sys.path:
    sys.path.insert(0, VENDOR_PATH)

try:
    from openpyxl import load_workbook
except ImportError as exc:
    print("Error: openpyxl is required. Install it with `python3 -m pip install --target scripts/_vendor openpyxl`.", file=sys.stderr)
    raise SystemExit(1) from exc


WORKBOOK_PATH = os.path.join(ROOT, "workbook", "rcmi_content.xlsx")
TARGETS = {
    "faculty": os.path.join(ROOT, "docs", "data", "faculty.csv"),
    "research": os.path.join(ROOT, "docs", "data", "research.csv"),
    "publications": os.path.join(ROOT, "docs", "data", "publications.csv"),
}


def fail(message):
    print(f"Error: {message}", file=sys.stderr)
    raise SystemExit(1)


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def read_sheet_rows(worksheet):
    rows = []

    # Find the maximum column to ensure all columns are included (including trailing empty ones)
    max_col = worksheet.max_column

    for row in worksheet.iter_rows(min_col=1, max_col=max_col, values_only=True):
        values = [normalize_cell(value) for value in row]
        rows.append(values)

    while rows and not any(rows[-1]):
        rows.pop()

    if not rows:
        return rows

    # Find the maximum width across all rows
    header_width = max(len(row) for row in rows)
    normalized = []

    for row in rows:
        # Ensure each row has exactly header_width columns
        if len(row) < header_width:
            row = row + [""] * (header_width - len(row))
        elif len(row) > header_width:
            row = row[:header_width]
        normalized.append(row)

    return normalized


def write_csv_atomic(path, rows):
    directory = os.path.dirname(path)
    os.makedirs(directory, exist_ok=True)
    fd, temp_path = tempfile.mkstemp(prefix=".tmp_", suffix=".csv", dir=directory)
    os.close(fd)
    try:
        with open(temp_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle, quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
            writer.writerows(rows)
        os.replace(temp_path, path)
    finally:
        if os.path.exists(temp_path):
            os.unlink(temp_path)


def export():
    if not os.path.exists(WORKBOOK_PATH):
        fail(f"Workbook file is missing: {WORKBOOK_PATH}")

    try:
        workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    except Exception as exc:
        fail(f"Unable to open workbook: {WORKBOOK_PATH} ({exc})")

    try:
        missing_sheets = [name for name in TARGETS if name not in workbook.sheetnames]
        if missing_sheets:
            fail("Workbook is missing required sheets: " + ", ".join(missing_sheets))

        summaries = []
        for sheet_name, target_path in TARGETS.items():
            worksheet = workbook[sheet_name]
            rows = read_sheet_rows(worksheet)
            if not rows:
                fail(f"Sheet '{sheet_name}' is empty in workbook: {WORKBOOK_PATH}")
            write_csv_atomic(target_path, rows)
            summaries.append((sheet_name, target_path, max(len(rows) - 1, 0)))
    finally:
        workbook.close()

    print("Regenerated CSV files from workbook:")
    for sheet_name, target_path, row_count in summaries:
        rel_path = os.path.relpath(target_path, ROOT)
        print(f"- {sheet_name} -> {rel_path} ({row_count} data rows)")


if __name__ == "__main__":
    export()
